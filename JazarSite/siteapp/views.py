from django.shortcuts import render, get_object_or_404
from google.auth.transport.requests import Request
from google.oauth2 import id_token
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
import openpyxl
from django.http import HttpResponse
from django.utils.encoding import smart_str
from django.contrib.auth.models import User
from rest_framework.parsers import MultiPartParser, FormParser
from django.core.mail import send_mail
from django.http import FileResponse, Http404
from django.conf import settings
import os
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.authentication import BaseAuthentication
from .models import Job, JobApplication, Blog, ContactUs
from .serializers import LoginSerializer, JobSerializer, JobApplicationSerializer, BlogSerializer, ContactUsSerializer
from rest_framework.permissions import IsAdminUser
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from django.http import JsonResponse
from rest_framework.decorators import api_view


class LoginAPIView(APIView):
    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        if serializer.is_valid():
            return Response(serializer.validated_data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class GoogleLoginAPIView(APIView):
    def post(self, request):
        # Get the google token from the frontend
        google_token = request.data.get("google_token")

        if not google_token:
            return Response({"error": "Google token is missing"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Verify the google token using Google's public keys
            idinfo = id_token.verify_oauth2_token(google_token, Request(), "YOUR_GOOGLE_CLIENT_ID")

            # Get user's email from token
            email = idinfo['email']

            # Check if the user exists
            user = User.objects.filter(email=email).first()
            if not user:
                user = User.objects.create_user(username=email, email=email)

            # Create JWT tokens
            refresh = RefreshToken.for_user(user)
            access_token = str(refresh.access_token)

            return Response({
                "access": access_token,
                "refresh": str(refresh),
                "is_admin": user.is_superuser,
                "message": "Google login successful"
            }, status=status.HTTP_200_OK)

        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class JobPostAPIView(APIView):
    def post(self, request):
        serializer = JobSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Job posted successfully!", "data": serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class JobListAPIView(APIView):
    def get(self, request):
        jobs = Job.objects.all().order_by('-created_at')  # Sort by newest jobs first
        serializer = JobSerializer(jobs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

class JobApplicationAPIView(APIView):
    def post(self, request):
        serializer = JobApplicationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Application submitted successfully!", "data": serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class BlogCreateAPIView(APIView):
    def post(self, request):
        serializer = BlogSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Blog post created successfully!", "data": serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class BlogListAPIView(APIView):
    def get(self, request):
        blogs = Blog.objects.all().order_by('-created_at')  # Sort by newest first
        serializer = BlogSerializer(blogs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class DeleteJobAPIView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def delete(self, request, job_id):
        try:
            job = Job.objects.get(id=job_id)
            job.delete()
            return Response({"detail": "Job deleted successfully."}, status=status.HTTP_200_OK)
        except Job.DoesNotExist:
            return Response({"detail": "Job not found."}, status=status.HTTP_404_NOT_FOUND)




class ContactUsAPIView(APIView):
    """
    API view to handle the submission of the contact form.
    It will save the data to the database and send an email.
    """
    
    def post(self, request):
        serializer = ContactUsSerializer(data=request.data)
        if serializer.is_valid():
            # Save the data into the database
            contact = serializer.save()

            # Send email to the admin
            subject = f"New message from {contact.first_name} {contact.last_name}"
            message = f"First Name: {contact.first_name}\nLast Name: {contact.last_name}\nContact Number: {contact.contact_number}\nEmail: {contact.email}\nMessage: {contact.message}"
            from_email = 'your_email@example.com'  # Replace with your email
            to_email = 'your_email@example.com'    # Replace with the email where you want to receive the form submissions

            send_mail(subject, message, from_email, [to_email])

            return Response({"message": "Your message has been sent successfully!"}, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



class UpdateJobAPIView(APIView):
    def put(self, request, job_id):
        try:
            job = Job.objects.get(id=job_id)
        except Job.DoesNotExist:
            return Response({"detail": "Job not found."}, status=status.HTTP_404_NOT_FOUND)

        serializer = JobSerializer(job, data=request.data, partial=True)

        if serializer.is_valid():
            serializer.save()
            return Response({"detail": "Job updated successfully."}, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



class JobApplicationListAPIView(APIView):
    """
    API view to fetch all job applications.
    """
    def get(self, request):
        try:
            applications = JobApplication.objects.all()  # Fetch all job applications
            serializer = JobApplicationSerializer(applications, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except JobApplication.DoesNotExist:
            return Response({"detail": "No applications found."}, status=status.HTTP_404_NOT_FOUND)



class ContactUsListAPIView(APIView):
    """
    API view to fetch a list of people who sent their details through the Contact Us form.
    """

    def get(self, request):
        try:
            # Fetch all contact submissions
            contact_us_entries = ContactUs.objects.all()
            serializer = ContactUsSerializer(contact_us_entries, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except ContactUs.DoesNotExist:
            return Response({"detail": "No contact entries found."}, status=status.HTTP_404_NOT_FOUND)


class DeleteBlogAPIView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def delete(self, request, blog_id):
        blog = get_object_or_404(Blog, id=blog_id)
        blog.delete()
        return Response({"detail": "Blog deleted successfully."}, status=status.HTTP_200_OK)

class UpdateBlogAPIView(APIView):
    parser_classes = (MultiPartParser, FormParser)  # Required for handling image uploads

    def put(self, request, blog_id):
        try:
            blog = Blog.objects.get(id=blog_id)
        except Blog.DoesNotExist:
            return Response({"detail": "Blog not found."}, status=status.HTTP_404_NOT_FOUND)

        serializer = BlogSerializer(blog, data=request.data, partial=True)

        if serializer.is_valid():
            serializer.save()
            return Response({"detail": "Blog updated successfully."}, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

from urllib.parse import unquote

def download_cv(request, user_id, filename):
    """Download CV files correctly from the media directory."""
    
    filename = unquote(filename)  # Decode URL-encoded filenames
    file_path = os.path.join(settings.MEDIA_ROOT, f'uploads/cv/{user_id}', filename)  # Adjusted path

    if os.path.exists(file_path):
        response = FileResponse(open(file_path, 'rb'), as_attachment=True)
        response['Content-Disposition'] = f'attachment; filename="{smart_str(filename)}"'
        return response
    else:
        raise Http404("File not found")
    

class JobApplicationExportView(APIView):
    def get(self, request):
        """Exports all job applications to an Excel file (without date filtering)."""

        # ✅ Create a new workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Job Applications"

        # ✅ Define headers
        headers = [
            "Full Name", "Email", "Phone Number", "Location",
            "Job Title", "Qualification", "Applied Date"
        ]
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header)

        # ✅ Query job applications
        queryset = JobApplication.objects.all().order_by("-applied_at")

        # ✅ Fill Excel rows with data
        row_index = 2
        for app in queryset:
            ws.cell(row=row_index, column=1, value=app.full_name)
            ws.cell(row=row_index, column=2, value=app.email)
            ws.cell(row=row_index, column=3, value=app.phone_number)
            ws.cell(row=row_index, column=4, value=app.location)
            ws.cell(row=row_index, column=5, value=app.job_title)
            ws.cell(row=row_index, column=6, value=app.qualification)
            ws.cell(row=row_index, column=7, value=app.applied_at.strftime("%Y-%m-%d %H:%M:%S"))
            row_index += 1

        # ✅ Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Job_Applications.xlsx'

        # ✅ Save the workbook to the response
        wb.save(response)
        return response
